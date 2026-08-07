"""Stage 5: filename safety, uniqueness, the Excel contract and the validator."""

import openpyxl
from PIL import Image

from src.models import Tile
from src.outputs import (EXCEL_COLUMNS, assign_image_names, slugify, validate,
                         write_excel)
from tests.helpers import img, size


def tile(name, spec, out_w=0, out_h=0, order=0):
    t = Tile(page_index=0, side="L", order=order,
             placement=img(0, 0, 100, 50), name=name, size=spec)
    t.out_w, t.out_h = out_w, out_h
    return t


# --- slugify ---------------------------------------------------------------

def test_slugify_strips_characters_windows_forbids():
    assert slugify('A/B:C*D?E"F|G<H>I') == "ABCDEFGHI"


def test_slugify_folds_whitespace_to_single_hyphens():
    assert slugify("ROLEX   BEIGE  HL") == "ROLEX-BEIGE-HL"
    assert slugify("EL_09__MATT") == "EL-09-MATT"


def test_slugify_keeps_a_readable_name():
    assert slugify("EL-09 (MATT)") == "EL-09-(MATT)"


def test_slugify_survives_a_missing_name():
    assert slugify("") == "UNNAMED"
    assert slugify(None) == "UNNAMED"


# --- assign_image_names ----------------------------------------------------

def test_image_names_carry_the_name_and_size():
    tiles = [tile("GOLD-02", size(1200, 600))]
    assign_image_names(tiles)
    assert tiles[0].image_name == "GOLD-02_1200x600.png"


def test_colliding_names_are_numbered_rather_than_overwritten():
    """Two faces of one mural, or the same product on two pages."""
    tiles = [tile("MONTANA-1004", size(500, 500)),
             tile("MONTANA-1004", size(500, 500)),
             tile("MONTANA-1004", size(500, 500))]
    assign_image_names(tiles)

    assert [t.image_name for t in tiles] == [
        "MONTANA-1004_500x500.png",
        "MONTANA-1004_500x500_2.png",
        "MONTANA-1004_500x500_3.png",
    ]


def test_the_same_name_at_a_different_size_does_not_collide():
    tiles = [tile("RS-1004-S", size(1000, 300)), tile("RS-1004-R", size(1000, 200))]
    assign_image_names(tiles)
    assert len(set(t.image_name for t in tiles)) == 2


# --- write_excel -----------------------------------------------------------

def test_the_workbook_has_exactly_the_four_required_columns(tmp_path):
    tiles = [tile("GOLD-02", size(1200, 600))]
    assign_image_names(tiles)
    path = tmp_path / "tiles.xlsx"

    write_excel(tiles, path)

    sheet = openpyxl.load_workbook(path).active
    assert [c.value for c in sheet[1]] == EXCEL_COLUMNS
    assert [c.value for c in sheet[2]] == ["GOLD-02", 1200, 600, "GOLD-02_1200x600.png"]


def test_length_column_always_holds_the_larger_millimetre_value(tmp_path):
    tiles = [tile("A", size(1200, 600)), tile("B", size(450, 300))]
    assign_image_names(tiles)
    path = tmp_path / "tiles.xlsx"

    write_excel(tiles, path)

    sheet = openpyxl.load_workbook(path).active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        assert row[1] >= row[2]


# --- validate --------------------------------------------------------------

def write_png(directory, name, w, h):
    Image.new("RGB", (w, h), "white").save(directory / name)


def test_a_clean_run_reports_no_problems(tmp_path):
    tiles = [tile("GOLD-02", size(1200, 600), out_w=1200, out_h=600)]
    assign_image_names(tiles)
    write_png(tmp_path, tiles[0].image_name, 1200, 600)

    assert validate(tiles, tmp_path) == []


def test_a_portrait_image_passes_against_its_landscape_millimetre_size(tmp_path):
    tiles = [tile("SUPER WHITE", size(1200, 600), out_w=606, out_h=1200)]
    assign_image_names(tiles)
    write_png(tmp_path, tiles[0].image_name, 606, 1200)

    assert validate(tiles, tmp_path) == []


def test_a_missing_title_is_reported(tmp_path):
    tiles = [tile("", size(600, 300), out_w=600, out_h=300)]
    assign_image_names(tiles)
    write_png(tmp_path, tiles[0].image_name, 600, 300)

    assert any("no title" in p for p in validate(tiles, tmp_path))


def test_a_missing_size_is_reported(tmp_path):
    tiles = [tile("A", None, out_w=600, out_h=300)]
    assign_image_names(tiles)
    write_png(tmp_path, tiles[0].image_name, 600, 300)

    assert any("non-positive dimensions" in p for p in validate(tiles, tmp_path))


def test_a_missing_image_file_is_reported(tmp_path):
    tiles = [tile("A", size(600, 300), out_w=600, out_h=300)]
    assign_image_names(tiles)

    assert any("missing image file" in p for p in validate(tiles, tmp_path))


def test_an_image_below_the_resolution_floor_is_reported(tmp_path):
    tiles = [tile("A", size(600, 300), out_w=300, out_h=150)]
    assign_image_names(tiles)
    write_png(tmp_path, tiles[0].image_name, 300, 150)

    assert any("below the 1px/mm floor" in p for p in validate(tiles, tmp_path))


def test_an_image_off_its_aspect_ratio_is_reported(tmp_path):
    tiles = [tile("A", size(600, 300), out_w=600, out_h=400)]
    assign_image_names(tiles)
    write_png(tmp_path, tiles[0].image_name, 600, 400)

    assert any("off their mm aspect ratio" in p for p in validate(tiles, tmp_path))


def test_duplicate_filenames_are_reported(tmp_path):
    tiles = [tile("A", size(600, 300), out_w=600, out_h=300),
             tile("A", size(600, 300), out_w=600, out_h=300)]
    for t in tiles:                       # bypass assign_image_names on purpose
        t.image_name = "A_600x300.png"
    write_png(tmp_path, "A_600x300.png", 600, 300)

    assert any("duplicate image filename" in p for p in validate(tiles, tmp_path))
